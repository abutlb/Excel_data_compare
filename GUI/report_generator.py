#!/usr/bin/env python
# -*- coding: utf-8 -*-

import os
import pandas as pd
import traceback
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def perform_comparison(app, comparison_column, output_file, main_file_info=None):
    """إجراء عملية المقارنة في خيط منفصل"""
    try:
        app.log("\n" + "=" * 50)
        app.log("بدء عملية المقارنة...")
        
        # تحضير البيانات للمقارنة
        dataframes = {}
        file_info = {}  # قاموس لتخزين معلومات الملفات
        error_files = []
        skipped_files = []
        
        # التعامل مع الملفات التي ليس لها أوراق محددة
        for file_path in app.excel_files:
            if "::" not in file_path and file_path not in app.selected_sheets:
                try:
                    # استخدام pandas للتحقق من وجود أوراق في الملف
                    xls = pd.ExcelFile(file_path)
                    sheet_names = xls.sheet_names
                    
                    if len(sheet_names) == 1:
                        # إذا كان الملف يحتوي على ورقة واحدة فقط، استخدمها تلقائياً
                        app.selected_sheets[file_path] = sheet_names[0]
                        app.log(f"تم تحديد الورقة الوحيدة '{sheet_names[0]}' تلقائياً للملف: {os.path.basename(file_path)}")
                    else:
                        app.log(f"تحذير: الملف '{os.path.basename(file_path)}' يحتوي على أكثر من ورقة ولم يتم تحديد ورقة. سيتم تخطيه.")
                        skipped_files.append({
                            'file': os.path.basename(file_path),
                            'reason': "لم يتم تحديد ورقة عمل",
                            'available_columns': ', '.join(sheet_names)
                        })
                except Exception as e:
                    app.log(f"خطأ في قراءة الملف '{os.path.basename(file_path)}': {str(e)}")
                    skipped_files.append({
                        'file': os.path.basename(file_path),
                        'reason': f"خطأ في قراءة الملف: {str(e)}",
                        'available_columns': '-'
                    })
        
        # قراءة جميع الملفات
        filtered_files = []
        for file_path in app.excel_files:
            if "::" in file_path or file_path in app.selected_sheets:
                filtered_files.append(file_path)
                
        total_files = len(filtered_files)
        app.log(f"عدد الملفات للمقارنة: {total_files}")
        
        for i, file_path in enumerate(filtered_files, 1):
            # تحديد اسم الملف والورقة
            if "::" in file_path:
                original_path = file_path.split("::")[0]
                sheet_name = file_path.split("::")[1]
                file_name = f"{os.path.basename(original_path)} - {sheet_name}"
                real_file_path = original_path
            else:
                file_name = os.path.basename(file_path)
                sheet_name = app.selected_sheets.get(file_path)
                real_file_path = file_path
            
            app.log(f"[{i}/{total_files}] جاري معالجة الملف: {file_name}...")
            app.log(f"قراءة الورقة: {sheet_name}")
            
            try:
                # قراءة البيانات من الورقة المحددة
                df = pd.read_excel(real_file_path, sheet_name=sheet_name)
                
                # تحديد العامود للمقارنة
                col = comparison_column
                if isinstance(comparison_column, dict):
                    col = comparison_column.get(file_path)
                
                app.log(f"عامود المقارنة المحدد: {col}")
                
                # التحقق من وجود العامود
                if col and col not in df.columns:
                    app.log(f"تحذير: العامود '{col}' غير موجود في الملف '{file_name}'. سيتم تخطي هذا الملف.")
                    app.log(f"الأعمدة المتاحة: {', '.join(df.columns)}")
                    skipped_files.append({
                        'file': file_name,
                        'reason': f"العامود '{col}' غير موجود",
                        'available_columns': ', '.join(df.columns)
                    })
                    continue
                
                # إزالة الصفوف التي تحتوي على قيم فارغة في عمود المقارنة
                if col:
                    original_count = len(df)
                    df = df.dropna(subset=[col])
                    if len(df) < original_count:
                        app.log(f"تم حذف {original_count - len(df)} صف يحتوي على قيم فارغة في عمود المقارنة")
                
                # تحديد نوع الملف (أساسي أو عادي)
                file_type = "عادي"
                if main_file_info and os.path.basename(main_file_info['path']) == os.path.basename(real_file_path):
                    file_type = "أساسي"
                
                # تخزين المعلومات
                dataframes[file_name] = df
                file_info[file_name] = {
                    'path': real_file_path,
                    'sheet': sheet_name,
                    'column': col,
                    'count': len(df),
                    'type': file_type,
                    'unique_count': 0  # سيتم تحديثه لاحقاً
                }
                
                app.log(f"تم قراءة الملف: {file_name} - عدد السجلات: {len(df)}")
                
            except Exception as e:
                app.log(f"خطأ في قراءة الملف '{file_name}': {str(e)}")
                error_files.append({
                    'file': file_name,
                    'error': str(e)
                })
        
        if not dataframes:
            app.log("لم يتم قراءة أي ملفات إكسل بنجاح.")
            app.finish_comparison(False)
            return
        
        # إنشاء قاموس يحتوي على جميع القيم الفريدة من جميع الملفات
        app.log("جاري تجميع القيم الفريدة من جميع الملفات...")
        
        # تحديد العمود الرئيسي للمقارنة
        main_column = None
        if isinstance(comparison_column, str):
            main_column = comparison_column
        elif main_file_info and main_file_info['path'] in comparison_column:
            main_column = comparison_column[main_file_info['path']]
        else:
            # استخدام العمود الأول من أول ملف
            first_file_name = list(dataframes.keys())[0]
            main_column = file_info[first_file_name]['column']
        
        app.log(f"العمود الرئيسي للمقارنة: {main_column}")
        
        # جمع جميع القيم الفريدة من جميع الملفات
        all_unique_values = set()
        for file_name, df in dataframes.items():
            column = file_info[file_name]['column']
            if column in df.columns:
                # تحويل القيم إلى نصوص للمقارنة الدقيقة
                values = set(df[column].dropna().astype(str).tolist())
                all_unique_values.update(values)
                app.log(f"تم جمع {len(values)} قيمة من الملف {file_name}")
        
        app.log(f"إجمالي عدد القيم الفريدة من جميع الملفات: {len(all_unique_values)}")
        
        # إنشاء DataFrame للمقارنة
        comparison_data = []
        for value in sorted(all_unique_values):
            row_data = {main_column: value}
            
            # إضافة عمود لكل ملف (1 إذا كانت القيمة موجودة، 0 إذا لم تكن موجودة)
            for file_name, df in dataframes.items():
                column = file_info[file_name]['column']
                if column in df.columns:
                    # التحقق من وجود القيمة في الملف
                    is_present = value in set(df[column].astype(str).tolist())
                    row_data[file_name] = 1 if is_present else 0
                else:
                    row_data[file_name] = 0
            
            comparison_data.append(row_data)
        
        # إنشاء DataFrame للمقارنة
        comparison_df = pd.DataFrame(comparison_data)
        
        # حساب عدد السجلات الفريدة لكل ملف
        for file_name in dataframes.keys():
            # السجلات الفريدة هي تلك الموجودة في هذا الملف فقط
            unique_count = sum(1 for row in comparison_data if row[file_name] == 1 and sum(row.get(other_file, 0) for other_file in dataframes.keys() if other_file != file_name) == 0)
            file_info[file_name]['unique_count'] = unique_count
            app.log(f"عدد السجلات الفريدة في الملف {file_name}: {unique_count}")
        
        # إنشاء ملف Excel للنتائج
        app.log("جاري إنشاء ملف Excel للنتائج...")
        
        # إنشاء ملف Excel باستخدام ExcelWriter
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # 1. إنشاء ورقة الملخص
            summary_data = []
            for file_name, info in file_info.items():
                summary_data.append({
                    'اسم الملف': file_name,
                    'عدد السجلات': info['count'],
                    'عدد السجلات الفريدة': info['unique_count'],
                    'عامود المقارنة': info['column'],
                    'ورقة العمل': info['sheet'],
                    'نوع الملف': info['type']
                })
            
            # إنشاء DataFrame للملخص
            summary_df = pd.DataFrame(summary_data)
            
            # إضافة صف الإجمالي
            total_records = sum(info['count'] for info in file_info.values())
            total_unique = len(all_unique_values)
            
            total_row = pd.DataFrame([{
                'اسم الملف': 'الإجمالي',
                'عدد السجلات': total_records,
                'عدد السجلات الفريدة': total_unique,
                'عامود المقارنة': '',
                'ورقة العمل': '',
                'نوع الملف': ''
            }])
            
            # دمج الملخص مع صف الإجمالي
            summary_df = pd.concat([summary_df, total_row], ignore_index=True)
            
            # كتابة ورقة الملخص
            summary_df.to_excel(writer, sheet_name='ملخص المقارنة', index=False)
            
            # 2. إنشاء ورقة للمقارنة
            comparison_df.to_excel(writer, sheet_name='مقارنة القيم', index=False)
        
        # تنسيق الملف بعد الكتابة
        app.log("جاري تنسيق ملف النتائج...")
        
        # فتح الملف للتنسيق
        wb = load_workbook(output_file)
        
        # تنسيق ورقة الملخص
        if 'ملخص المقارنة' in wb:
            ws = wb['ملخص المقارنة']
            ws.sheet_view.rightToLeft = True  # تعيين اتجاه الورقة من اليمين إلى اليسار
            
            # تنسيق العناوين
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            # تنسيق الصف الأول (العناوين)
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # تنسيق صف الإجمالي
            total_row = len(summary_data) + 1
            for cell in ws[total_row + 1]:  # +1 لأن الصفوف تبدأ من 1 وليس 0
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D8E4BC", end_color="D8E4BC", fill_type="solid")
        
        # تنسيق ورقة المقارنة
        if 'مقارنة القيم' in wb:
            ws = wb['مقارنة القيم']
            ws.sheet_view.rightToLeft = True
            
            # تنسيق العناوين
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # تنسيق ورقة الملفات المتخطاة
        if 'الملفات المتخطاة' in wb:
            ws = wb['الملفات المتخطاة']
            ws.sheet_view.rightToLeft = True
            
            # تنسيق العناوين
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # حفظ الملف بعد التنسيق
        wb.save(output_file)
        app.log(f"تم حفظ النتائج في الملف: {output_file}")
        app.log("تم حفظ النتائج بنجاح.")
        
        # إنهاء المقارنة
        app.finish_comparison(True)
        
    except Exception as e:
        app.log(f"حدث خطأ أثناء المقارنة: {str(e)}")
        app.log(traceback.format_exc())
        app.finish_comparison(False)