#!/usr/bin/env python
# -*- coding: utf-8 -*-

import os
import pandas as pd
import argparse
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Color
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

def compare_excel_files(input_folder='in', comparison_column=None, output_file='comparison_results.xlsx'):
    """
    قارن بين ملفات الإكسل في المجلد المحدد بناءً على العامود المحدد
    
    المعاملات:
    input_folder (str): مسار المجلد الذي يحتوي على ملفات الإكسل
    comparison_column (str): اسم العامود الذي سيتم المقارنة على أساسه
    output_file (str): اسم ملف الإكسل الناتج
    """
    # التحقق من وجود المجلد
    if not os.path.exists(input_folder):
        os.makedirs(input_folder)
        print(f"تم إنشاء مجلد '{input_folder}'. يرجى وضع ملفات الإكسل فيه وإعادة تشغيل السكربت.")
        return
    
    # الحصول على قائمة ملفات الإكسل في المجلد
    excel_files = [f for f in os.listdir(input_folder) if f.endswith(('.xlsx', '.xls'))]
    
    if not excel_files:
        print(f"لم يتم العثور على ملفات إكسل في المجلد '{input_folder}'.")
        return
    
    # قراءة جميع ملفات الإكسل
    dataframes = {}
    all_data = []
    error_files = []  # ملفات بها أخطاء
    skipped_files = []  # ملفات تم تخطيها لعدم وجود العامود المحدد
    
    for file in excel_files:
        file_path = os.path.join(input_folder, file)
        try:
            df = pd.read_excel(file_path)
            
            # التحقق من وجود العامود المحدد للمقارنة
            if comparison_column and comparison_column not in df.columns:
                print(f"تحذير: العامود '{comparison_column}' غير موجود في الملف '{file}'. سيتم تخطي هذا الملف.")
                print(f"الأعمدة المتاحة في الملف '{file}': {', '.join(df.columns)}")
                skipped_files.append({
                    'file': file,
                    'reason': f"العامود '{comparison_column}' غير موجود",
                    'available_columns': ', '.join(df.columns)
                })
                continue
                
            # إضافة عامود يحدد مصدر البيانات (اسم الملف)
            df['source_file'] = file
            
            dataframes[file] = df
            all_data.append(df)
            print(f"تم قراءة الملف: {file} - عدد السجلات: {len(df)}")
        except Exception as e:
            print(f"خطأ في قراءة الملف '{file}': {str(e)}")
            error_files.append({
                'file': file,
                'error': str(e)
            })
    
    if not dataframes:
        print("لم يتم قراءة أي ملفات إكسل بنجاح.")
        return
    
    # دمج جميع البيانات
    combined_df = pd.concat(all_data, ignore_index=True)
    
    # إنشاء مجموعة فريدة من السجلات بناءً على العامود المحدد
    if comparison_column:
        unique_records = combined_df.drop_duplicates(subset=[comparison_column])
    else:
        # إذا لم يتم تحديد عامود، استخدم جميع الأعمدة باستثناء source_file
        columns_to_check = [col for col in combined_df.columns if col != 'source_file']
        unique_records = combined_df.drop_duplicates(subset=columns_to_check)
    
    # إنشاء ملف إكسل جديد
    wb = Workbook()
    
    # تنسيق الخلايا
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    centered_alignment = Alignment(horizontal='center', vertical='center')
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # تنسيقات إضافية
    success_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    warning_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    error_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    success_font = Font(color="006100")
    warning_font = Font(color="9C5700")
    error_font = Font(color="9C0006")
    
    # إنشاء ورقة للسجلات الفريدة
    ws_unique = wb.active
    ws_unique.title = "جميع السجلات الفريدة"
    
    # إضافة عنوان للورقة
    ws_unique.merge_cells('A1:E1')
    title_cell = ws_unique.cell(row=1, column=1, value="جميع السجلات الفريدة (بدون تكرار)")
    title_cell.font = Font(bold=True, size=14, color="366092")
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # إضافة وصف
    if comparison_column:
        description = f"السجلات الفريدة بناءً على العامود: {comparison_column}"
    else:
        description = "السجلات الفريدة بناءً على جميع الأعمدة"
    
    ws_unique.merge_cells('A2:E2')
    desc_cell = ws_unique.cell(row=2, column=1, value=description)
    desc_cell.font = Font(italic=True)
    desc_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # إضافة البيانات الفريدة
    start_row = 4
    for r_idx, row in enumerate(dataframe_to_rows(unique_records, index=False, header=True), start_row):
        for c_idx, value in enumerate(row, 1):
            cell = ws_unique.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == start_row:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = centered_alignment
            cell.border = border
    
    # تعديل عرض الأعمدة
    for column in ws_unique.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws_unique.column_dimensions[column_letter].width = adjusted_width
    
    # إنشاء ورقة لكل ملف تعرض السجلات الفريدة له
    for file_name, df in dataframes.items():
        sheet_name = os.path.splitext(file_name)[0]
        if len(sheet_name) > 31:  # حد أقصى لاسم الورقة في إكسل
            sheet_name = sheet_name[:31]
        
        ws_file = wb.create_sheet(title=sheet_name)
        
        # إضافة عنوان للورقة
        ws_file.merge_cells('A1:E1')
        title_cell = ws_file.cell(row=1, column=1, value=f"السجلات الفريدة في الملف: {file_name}")
        title_cell.font = Font(bold=True, size=14, color="366092")
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # إضافة وصف
        if comparison_column:
            description = f"السجلات الموجودة في هذا الملف فقط (بناءً على العامود: {comparison_column})"
        else:
            description = "السجلات الموجودة في هذا الملف فقط (بناءً على جميع الأعمدة)"
        
        ws_file.merge_cells('A2:E2')
        desc_cell = ws_file.cell(row=2, column=1, value=description)
        desc_cell.font = Font(italic=True)
        desc_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # البحث عن السجلات الفريدة لهذا الملف
        unique_to_file = []
        
        if comparison_column:
            # الحصول على قيم العامود المحدد من الملف الحالي
            current_values = set(df[comparison_column].tolist())
            
            # الحصول على قيم العامود المحدد من الملفات الأخرى
            other_values = set()
            for other_file, other_df in dataframes.items():
                if other_file != file_name:
                    other_values.update(other_df[comparison_column].tolist())
            
            # السجلات الفريدة هي تلك الموجودة في الملف الحالي ولكن ليست في الملفات الأخرى
            unique_values = current_values - other_values
            unique_to_file = df[df[comparison_column].isin(unique_values)]
        else:
            # إذا لم يتم تحديد عامود، قم بمقارنة جميع الأعمدة باستثناء source_file
            columns_to_check = [col for col in df.columns if col != 'source_file']
            
            for _, row in df.iterrows():
                is_unique = True
                for other_file, other_df in dataframes.items():
                    if other_file != file_name:
                        # التحقق مما إذا كانت هذه السجلة موجودة في الملف الآخر
                        match_mask = True
                        for col in columns_to_check:
                            match_mask = match_mask & (other_df[col] == row[col])
                        
                        if match_mask.any():
                            is_unique = False
                            break
                
                if is_unique:
                    unique_to_file.append(row)
            
            if unique_to_file:
                unique_to_file = pd.DataFrame(unique_to_file)
        
        # إضافة معلومات إحصائية
        ws_file.cell(row=4, column=1, value="إجمالي عدد السجلات في الملف:").font = Font(bold=True)
        ws_file.cell(row=4, column=2, value=len(df))
        
        ws_file.cell(row=5, column=1, value="عدد السجلات الفريدة:").font = Font(bold=True)
        ws_file.cell(row=5, column=2, value=len(unique_to_file) if isinstance(unique_to_file, pd.DataFrame) else len(unique_to_file))
        
        # إضافة البيانات الفريدة للملف
        start_row = 7
        if isinstance(unique_to_file, pd.DataFrame) and len(unique_to_file) > 0:
            for r_idx, row in enumerate(dataframe_to_rows(unique_to_file, index=False, header=True), start_row):
                for c_idx, value in enumerate(row, 1):
                    cell = ws_file.cell(row=r_idx, column=c_idx, value=value)
                    if r_idx == start_row:
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = centered_alignment
                    cell.border = border
            
            # تعديل عرض الأعمدة
            for column in ws_file.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws_file.column_dimensions[column_letter].width = adjusted_width
        else:
            # إذا لم تكن هناك سجلات فريدة، أضف رسالة
            ws_file.merge_cells(f'A{start_row}:E{start_row}')
            no_data_cell = ws_file.cell(row=start_row, column=1, value="لا توجد سجلات فريدة لهذا الملف")
            no_data_cell.font = Font(italic=True, color="9C0006")
            no_data_cell.alignment = centered_alignment
    
    # إنشاء ورقة ملخص
    ws_summary = wb.create_sheet(title="ملخص المقارنة", index=0)
    
    # إضافة معلومات الملخص
    ws_summary.merge_cells('A1:F1')
    title_cell = ws_summary.cell(row=1, column=1, value="ملخص مقارنة ملفات الإكسل")
    title_cell.font = Font(bold=True, size=16, color="366092")
    title_cell.alignment = centered_alignment
    
    # إضافة تاريخ التقرير
    from datetime import datetime
    ws_summary.cell(row=2, column=1, value=f"تاريخ التقرير: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    ws_summary.cell(row=2, column=1).font = Font(italic=True)
    
    # إضافة معلومات المقارنة
    ws_summary.cell(row=4, column=1, value="معلومات المقارنة").font = Font(bold=True, size=12, color="366092")
    
    ws_summary.cell(row=5, column=1, value="عدد الملفات التي تم العثور عليها:").font = Font(bold=True)
    ws_summary.cell(row=5, column=2, value=len(excel_files))
    
    ws_summary.cell(row=6, column=1, value="عدد الملفات التي تمت قراءتها بنجاح:").font = Font(bold=True)
    ws_summary.cell(row=6, column=2, value=len(dataframes))
    ws_summary.cell(row=6, column=2).font = success_font
    ws_summary.cell(row=6, column=2).fill = success_fill
    
    ws_summary.cell(row=7, column=1, value="عدد الملفات التي تم تخطيها:").font = Font(bold=True)
    ws_summary.cell(row=7, column=2, value=len(skipped_files))
    if skipped_files:
        ws_summary.cell(row=7, column=2).font = warning_font
        ws_summary.cell(row=7, column=2).fill = warning_fill
    
    ws_summary.cell(row=8, column=1, value="عدد الملفات التي بها أخطاء:").font = Font(bold=True)
    ws_summary.cell(row=8, column=2, value=len(error_files))
    if error_files:
        ws_summary.cell(row=8, column=2).font = error_font
        ws_summary.cell(row=8, column=2).fill = error_fill
    
    ws_summary.cell(row=10, column=1, value="إجمالي عدد السجلات:").font = Font(bold=True)
    ws_summary.cell(row=10, column=2, value=len(combined_df))
    
    ws_summary.cell(row=11, column=1, value="عدد السجلات الفريدة:").font = Font(bold=True)
    ws_summary.cell(row=11, column=2, value=len(unique_records))
    
    if comparison_column:
        ws_summary.cell(row=12, column=1, value="العامود المستخدم للمقارنة:").font = Font(bold=True)
        ws_summary.cell(row=12, column=2, value=comparison_column)
    else:
        ws_summary.cell(row=12, column=1, value="طريقة المقارنة:").font = Font(bold=True)
        ws_summary.cell(row=12, column=2, value="مقارنة جميع الأعمدة")
    
    # إضافة جدول تفاصيل الملفات
    ws_summary.cell(row=14, column=1, value="تفاصيل الملفات التي تمت قراءتها بنجاح").font = Font(bold=True, size=12, color="366092")
    
    # إضافة رأس الجدول
    headers = ["اسم الملف", "عدد السجلات", "عدد السجلات الفريدة", "النسبة المئوية للسجلات الفريدة"]
    for i, header in enumerate(headers, 1):
        cell = ws_summary.cell(row=15, column=i, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = centered_alignment
        cell.border = border
    
    # إضافة بيانات الملفات
    row_idx = 16
    for file_name, df in dataframes.items():
        # حساب عدد السجلات الفريدة لهذا الملف
        if comparison_column:
            current_values = set(df[comparison_column].tolist())
            other_values = set()
            for other_file, other_df in dataframes.items():
                if other_file != file_name:
                    other_values.update(other_df[comparison_column].tolist())
            unique_values = current_values - other_values
            unique_count = len(unique_values)
        else:
            # هذا تقريبي لأن الحساب الدقيق يتطلب مقارنة كل سجل
            unique_count = len(df.drop_duplicates(subset=[col for col in df.columns if col != 'source_file']))
            for other_file, other_df in dataframes.items():
                if other_file != file_name:
                    unique_count -= len(df.merge(other_df, how='inner', on=[col for col in df.columns if col != 'source_file']))
            unique_count = max(0, unique_count)
        
        # حساب النسبة المئوية
        percentage = (unique_count / len(df)) * 100 if len(df) > 0 else 0
        
        # إضافة بيانات الملف
        ws_summary.cell(row=row_idx, column=1, value=file_name).border = border
        ws_summary.cell(row=row_idx, column=2, value=len(df)).border = border
        ws_summary.cell(row=row_idx, column=3, value=unique_count).border = border
        ws_summary.cell(row=row_idx, column=4, value=f"{percentage:.2f}%").border = border
        
        row_idx += 1
    
    # إضافة رسم بياني للسجلات
    chart = BarChart()
    chart.title = "مقارنة عدد السجلات بين الملفات"
    chart.style = 10
    chart.x_axis.title = "الملفات"
    chart.y_axis.title = "عدد السجلات"
    
    data = Reference(ws_summary, min_col=2, min_row=15, max_row=row_idx-1, max_col=3)
    cats = Reference(ws_summary, min_col=1, min_row=16, max_row=row_idx-1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.shape = 4
    ws_summary.add_chart(chart, "H15")
    
    # إضافة جدول للملفات التي تم تخطيها
    if skipped_files:
        start_row = row_idx + 2
        ws_summary.cell(row=start_row, column=1, value="الملفات التي تم تخطيها").font = Font(bold=True, size=12, color="9C5700")
        
        # إضافة رأس الجدول
        headers = ["اسم الملف", "سبب التخطي", "الأعمدة المتاحة"]
        for i, header in enumerate(headers, 1):
            cell = ws_summary.cell(row=start_row+1, column=i, value=header)
            cell.font = header_font
            cell.fill = warning_fill
            cell.alignment = centered_alignment
            cell.border = border
            cell.font = Font(bold=True, color="9C5700")
        
        # إضافة بيانات الملفات التي تم تخطيها
        for i, file_info in enumerate(skipped_files):
            row = start_row + 2 + i
            ws_summary.cell(row=row, column=1, value=file_info['file']).border = border
            ws_summary.cell(row=row, column=2, value=file_info['reason']).border = border
            ws_summary.cell(row=row, column=3, value=file_info['available_columns']).border = border
        
        row_idx = start_row + 2 + len(skipped_files)
    
    # إضافة جدول للملفات التي بها أخطاء
    if error_files:
        start_row = row_idx + 2
        ws_summary.cell(row=start_row, column=1, value="الملفات التي بها أخطاء").font = Font(bold=True, size=12, color="9C0006")
        
        # إضافة رأس الجدول
        headers = ["اسم الملف", "وصف الخطأ"]
        for i, header in enumerate(headers, 1):
            cell = ws_summary.cell(row=start_row+1, column=i, value=header)
            cell.font = header_font
            cell.fill = error_fill
            cell.alignment = centered_alignment
            cell.border = border
            cell.font = Font(bold=True, color="9C0006")
        
        # إضافة بيانات الملفات التي بها أخطاء
        for i, file_info in enumerate(error_files):
            row = start_row + 2 + i
            ws_summary.cell(row=row, column=1, value=file_info['file']).border = border
            ws_summary.cell(row=row, column=2, value=file_info['error']).border = border
    
    # تعديل عرض الأعمدة في ورقة الملخص
    for column in ws_summary.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws_summary.column_dimensions[column_letter].width = adjusted_width
    
    # حفظ الملف
    try:
        wb.save(output_file)
        print(f"تم إنشاء ملف المقارنة بنجاح: {output_file}")
        return output_file
    except Exception as e:
        print(f"خطأ في حفظ الملف: {str(e)}")
        # محاولة الحفظ باسم مختلف في حالة كان الملف مفتوحاً
        alternative_file = f"comparison_results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        try:
            wb.save(alternative_file)
            print(f"تم حفظ الملف باسم بديل: {alternative_file}")
            return alternative_file
        except Exception as e2:
            print(f"فشل في حفظ الملف البديل: {str(e2)}")
            return None

def main():
    # إعداد محلل وسائط سطر الأوامر
    parser = argparse.ArgumentParser(description="برنامج مقارنة ملفات الإكسل")
    parser.add_argument("--input", "-i", help="مجلد يحتوي على ملفات الإكسل للمقارنة", default="in")
    parser.add_argument("--column", "-c", help="اسم العامود الذي سيتم المقارنة على أساسه")
    parser.add_argument("--output", "-o", help="اسم ملف الإكسل الناتج", default="comparison_results.xlsx")
    parser.add_argument("--interactive", "-int", action="store_true", help="تشغيل البرنامج في الوضع التفاعلي")
    
    # تحليل وسائط سطر الأوامر
    args = parser.parse_args()
    
    # تحديد ما إذا كان يجب تشغيل البرنامج في الوضع التفاعلي
    if args.interactive:
        run_interactive_mode()
    else:
        # تشغيل البرنامج باستخدام الوسائط المحددة
        input_folder = args.input
        comparison_column = args.column
        output_file = args.output
        
        # التأكد من وجود لاحقة .xlsx في اسم ملف الإخراج
        if output_file and not output_file.endswith('.xlsx'):
            output_file += '.xlsx'
        
        print("\n" + "=" * 60)
        print("برنامج مقارنة ملفات الإكسل".center(60))
        print("=" * 60)
        
        print(f"\nالمجلد المحدد: {input_folder}")
        print(f"العامود المستخدم للمقارنة: {comparison_column if comparison_column else 'جميع الأعمدة'}")
        print(f"ملف الإخراج: {output_file}")
        print("\n" + "-" * 60)
        
        # تنفيذ المقارنة
        result_file = compare_excel_files(
            input_folder=input_folder,
            comparison_column=comparison_column,
            output_file=output_file
        )
        
        if result_file:
            print("\n" + "=" * 60)
            print(f"تمت العملية بنجاح! تم إنشاء ملف: {result_file}")
            
            # محاولة فتح الملف تلقائياً
            try:
                import os
                os.system(f'start excel "{result_file}"')
                print("تم فتح الملف تلقائياً.")
            except:
                print(f"يمكنك فتح الملف يدوياً من: {os.path.abspath(result_file)}")
        
        print("\n" + "=" * 60)

def run_interactive_mode():
    """تشغيل البرنامج في الوضع التفاعلي حيث يطلب من المستخدم إدخال المعلومات"""
    print("\n" + "=" * 60)
    print("برنامج مقارنة ملفات الإكسل".center(60))
    print("=" * 60)
    print("\nيرجى التأكد من وضع جميع ملفات الإكسل في مجلد 'in'")
    
    # طلب مجلد الإدخال
    input_folder = input("\nأدخل مسار مجلد الإدخال (الافتراضي: 'in'): ")
    if input_folder.strip() == "":
        input_folder = "in"
    
    # التحقق من وجود المجلد وإنشائه إذا لم يكن موجوداً
    if not os.path.exists(input_folder):
        os.makedirs(input_folder)
        print(f"\nتم إنشاء مجلد '{input_folder}'. يرجى وضع ملفات الإكسل فيه وإعادة تشغيل البرنامج.")
        input("\nاضغط Enter للخروج...")
        exit()
    
    # التحقق من وجود ملفات إكسل في المجلد
    excel_files = [f for f in os.listdir(input_folder) if f.endswith(('.xlsx', '.xls'))]
    if not excel_files:
        print(f"\nلم يتم العثور على ملفات إكسل في مجلد '{input_folder}'.")
        print("يرجى وضع ملفات الإكسل في المجلد وإعادة تشغيل البرنامج.")
        input("\nاضغط Enter للخروج...")
        exit()
    
    print(f"\nتم العثور على {len(excel_files)} ملف إكسل في مجلد '{input_folder}':")
    for i, file in enumerate(excel_files, 1):
        print(f"{i}. {file}")
    
    # محاولة قراءة أول ملف للحصول على قائمة الأعمدة المتاحة
    try:
        first_file = os.path.join(input_folder, excel_files[0])
        df = pd.read_excel(first_file)
        available_columns = df.columns.tolist()
        
        print("\nالأعمدة المتاحة في الملف الأول:")
        for i, col in enumerate(available_columns, 1):
            print(f"{i}. {col}")
        
        print("\nملاحظة: قد تختلف الأعمدة المتاحة بين الملفات.")
    except Exception as e:
        print(f"\nتعذر قراءة الملف الأول: {str(e)}")
        available_columns = []
    
    # طلب اسم العامود للمقارنة
    print("\n" + "-" * 60)
    comparison_column = input("أدخل اسم العامود الذي تريد المقارنة على أساسه (اضغط Enter للمقارنة على أساس جميع الأعمدة): ")
    if comparison_column.strip() == "":
        comparison_column = None
        print("سيتم المقارنة على أساس جميع الأعمدة.")
    else:
        print(f"سيتم المقارنة على أساس العامود: {comparison_column}")
    
    # طلب اسم ملف الإخراج
    output_file = input("\nأدخل اسم ملف الإخراج (الافتراضي: comparison_results.xlsx): ")
    if output_file.strip() == "":
        output_file = "comparison_results.xlsx"
    
    # إضافة لاحقة .xlsx إذا لم تكن موجودة
    if not output_file.endswith('.xlsx'):
        output_file += '.xlsx'
    
    print("\n" + "-" * 60)
    print("جاري معالجة الملفات...")
    
    # تنفيذ المقارنة
    result_file = compare_excel_files(input_folder=input_folder, comparison_column=comparison_column, output_file=output_file)
    
    if result_file:
        print("\n" + "=" * 60)
        print(f"تمت العملية بنجاح! تم إنشاء ملف: {result_file}")
        
        # محاولة فتح الملف تلقائياً
        try:
            import os
            os.system(f'start excel "{result_file}"')
            print("تم فتح الملف تلقائياً.")
        except:
            print(f"يمكنك فتح الملف يدوياً من: {os.path.abspath(result_file)}")
    
    print("\n" + "=" * 60)
    input("اضغط Enter للخروج...")

if __name__ == "__main__":
    main()